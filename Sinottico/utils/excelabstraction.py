import sys
import datetime
import os

#import logging
#logging.basicConfig(filename='sinottico.log', level=logging.INFO)


class CustomExcelWorkbookBecauseWindowsSucks:
    def _openpyxl_init(self, filename):
        self.native = False
        from openpyxl import Workbook, load_workbook
        self.wb = load_workbook(filename)
        self.ws = self.wb.active

        if self.excel:
            self.excel.quit()

    def _open_excel(self, filename):
        try:
            import win32com.client
            import pywintypes
            try:
                self.native = True
                self.excel = win32com.client.Dispatch('Excel.Application')
                self.wb = self.excel.Workbooks.Open(filename)
                self.ws = self.wb.Worksheets(1)
                #print("Caricato il modulo Excel Nativo")
            except pywintypes.com_error as e:
                #print("Excel non installato, procedo con openpyxl")
                return False
        except ModuleNotFoundError:
            #print("win32com non installato, procedo con openpyxl")
            return False
        return True

    def __init__(self, filename):
        self.excel = None
        filename = os.path.abspath(filename)
        # print(filename)
        import platform
        if platform.system() in ["Windows", "windows"]:
            if not self._open_excel(filename):
                if self.excel != None:
                    self.excel.quit()
                    if self._open_excel(filename):
                        return
                self._openpyxl_init(filename)
        else:
            #print("Sistema operativo superiore, procedo con openpyxl")
            self._openpyxl_init(filename)

    def __setitem__(self, id, value):
        if self.native:
            self.ws.Range(id).Value = value
        else:
            self.ws[id] = value


    def __getitem__(self, id, value):
        if self.native:
            return self.ws.Range(id).Value
        else:
            return self.ws[id]


    def __del__(self):
        if self.native and self.excel:
            self.excel.quit()


    def save(self, filename):
        if self.native:
            self.wb.SaveAs(os.path.abspath(filename))
        else:
            self.wb.save(filename=filename)


    def cell_id(self, x, y): 
        return "{}{}".format(chr(y), x)


    def read_data(self):
        data = DatiPotenza()

        for t in [25, 45]:
            for a in range(1, 33):
                data.diretta[t][a] = (self[self.index_to_cell_diretta(t, a, 0)],
                                        self[self.index_to_cell_diretta(t, a, 1)],
                                        self[self.index_to_cell_diretta(t, a, 2)])

        for t in [25, 45]:
            for a in range(1, 33):
                data.riflessa[t][a] = (self[self.index_to_cell_riflessa(t, a, 0)],
                                        self[self.index_to_cell_riflessa(t, a, 1)],
                                        self[self.index_to_cell_riflessa(t, a, 2)])

        return data


    def write_data(self, data):
        for t in [25, 45]:
            for a in data.diretta[t].keys():
                self[self.index_to_cell_diretta_base(a)] = a
                self[self.index_to_cell_diretta(t,a,0)] = data.diretta[t][a][0]
                self[self.index_to_cell_diretta(t,a,1)] = data.diretta[t][a][1]
                self[self.index_to_cell_diretta(t,a,2)] = data.diretta[t][a][2]

        for t in [25, 45]:
            for a in data.riflessa[t].keys():
                self[self.index_to_cell_riflessa_base(a)] = a
                self[self.index_to_cell_riflessa(t,a,0)] = data.riflessa[t][a][0]
                self[self.index_to_cell_riflessa(t,a,1)] = data.riflessa[t][a][1]
                self[self.index_to_cell_riflessa(t,a,2)] = data.riflessa[t][a][2]


    def index_to_cell_riflessa(self, temp : int, att : int, index : int):
        scol = {25: 66, 45: 72}[temp] + index
        if att > 32:
            att = 32
        return self.cell_id(47 + 32 - att, scol)


    def index_to_cell_diretta(self, temp : int, att : int, index : int):
        scol = {25: 66, 45: 70}[temp] + index
        if att > 32:
            att = 32
        return self.cell_id(11 + 32 - att, scol)

    def index_to_cell_diretta_base(self, att : int):
        return self.cell_id(11 + 32 - att, 65)

    def index_to_cell_riflessa_base(self, att : int):
        return self.cell_id(47 + 32 - att, 65)